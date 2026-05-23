# reports/excel_report.py

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import warnings
import time
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment,
    Border, Side, GradientFill,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

from src.config import (
    REPORTS_DIR,
    create_directories,
)
from src.sql_queries import (
    get_business_overview,
    get_monthly_revenue,
    get_quarterly_revenue,
    get_revenue_by_category,
    get_revenue_by_state,
    get_customer_segmentation,
    get_order_status_summary,
    get_payment_analysis,
    get_review_analysis,
    get_delivery_analysis,
    get_seller_performance,
    get_rfm_segment_summary,
    get_top_customers,
    get_top_products,
    get_running_total_revenue,
    get_category_performance,
    get_orders_by_weekday,
    get_orders_by_hour,
)

warnings.filterwarnings("ignore")

# ── Configure Logger ──────────────────────────────────────────────────────
logger.remove()
logger.add(
    sys.stdout,
    format="{time:YYYY-MM-DD HH:mm:ss} | {level: <8} | {name} | {message}",
    level="INFO",
    colorize=False,
)

# ── Color Constants ───────────────────────────────────────────────────────
DARK_BG     = "0F1117"
CARD_BG     = "1A1D27"
BLUE        = "3498DB"
GREEN       = "2ECC71"
RED         = "E74C3C"
ORANGE      = "F39C12"
PURPLE      = "9B59B6"
HEADER_BG   = "2C3E50"
ALT_ROW     = "F5F6FA"
WHITE       = "FFFFFF"
LIGHT_BLUE  = "EBF5FB"
LIGHT_GREEN = "EAFAF1"
LIGHT_RED   = "FDEDEC"
LIGHT_ORANGE= "FEF9E7"
BORDER_CLR  = "BDC3C7"


# ════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ════════════════════════════════════════════════════════════

def make_header_fill(color: str) -> PatternFill:
    return PatternFill(
        start_color=color,
        end_color=color,
        fill_type="solid"
    )


def make_border() -> Border:
    thin = Side(style="thin", color=BORDER_CLR)
    return Border(
        left=thin, right=thin,
        top=thin, bottom=thin
    )


def make_thick_border() -> Border:
    thick = Side(style="medium", color=HEADER_BG)
    return Border(
        left=thick, right=thick,
        top=thick, bottom=thick
    )


def style_header_row(
    ws,
    row: int,
    start_col: int,
    end_col: int,
    bg_color: str = HEADER_BG,
    font_color: str = WHITE,
    font_size: int = 10,
) -> None:
    """
    Apply header styling to a row.
    """
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = make_header_fill(bg_color)
        cell.font      = Font(
            bold=True, color=font_color,
            size=font_size, name="Arial"
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border    = make_border()


def style_data_row(
    ws,
    row: int,
    start_col: int,
    end_col: int,
    alt_row: bool = False,
) -> None:
    """
    Apply data row styling.
    """
    bg = ALT_ROW if alt_row else WHITE
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = make_header_fill(bg)
        cell.font      = Font(size=9, name="Arial")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border    = make_border()


def set_column_widths(
    ws,
    widths: list
) -> None:
    """
    Set column widths for a worksheet.
    """
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[
            get_column_letter(i)
        ].width = width


def write_dataframe(
    ws,
    df: pd.DataFrame,
    start_row: int = 1,
    start_col: int = 1,
    header_color: str = HEADER_BG,
) -> int:
    """
    Write a DataFrame to a worksheet with styling.
    Returns the last row written.
    """
    n_cols = len(df.columns)

    # Write header
    for j, col in enumerate(df.columns, start_col):
        cell       = ws.cell(row=start_row, column=j)
        cell.value = str(col).replace("_", " ").title()

    style_header_row(
        ws, start_row, start_col,
        start_col + n_cols - 1, header_color
    )

    # Write data
    for i, (_, row_data) in enumerate(df.iterrows()):
        data_row  = start_row + i + 1
        alt       = i % 2 == 1
        for j, val in enumerate(row_data, start_col):
            cell       = ws.cell(row=data_row, column=j)
            cell.value = val if not pd.isna(val) else ""
            style_data_row(
                ws, data_row, start_col,
                start_col + n_cols - 1, alt
            )

    return start_row + len(df) + 1


def make_title(
    ws,
    title: str,
    row: int,
    n_cols: int,
    color: str = BLUE,
) -> None:
    """
    Write a merged title cell.
    """
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=n_cols
    )
    cell           = ws.cell(row=row, column=1)
    cell.value     = title
    cell.font      = Font(
        bold=True, size=14,
        color=color, name="Arial"
    )
    cell.alignment = Alignment(
        horizontal="center", vertical="center"
    )
    cell.fill      = make_header_fill(LIGHT_BLUE)
    ws.row_dimensions[row].height = 30


# ════════════════════════════════════════════════════════════
# SHEET 1: OVERVIEW
# ════════════════════════════════════════════════════════════

def write_overview_sheet(
    wb: Workbook,
    data: dict
) -> None:
    """
    Write business overview KPI sheet.
    """
    logger.info("-- Writing Overview sheet --")

    ws         = wb.active
    ws.title   = "Overview"
    ws.sheet_view.showGridLines = False

    ov = data["overview"].iloc[0]

    # Title
    make_title(
        ws, "RETAIL ANALYTICS - BUSINESS OVERVIEW",
        1, 6, BLUE
    )

    # Subtitle
    ws.merge_cells("A2:F2")
    cell           = ws["A2"]
    cell.value     = (
        f"Brazilian E-Commerce (Olist) | "
        f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    )
    cell.font      = Font(
        size=10, color=HEADER_BG,
        italic=True, name="Arial"
    )
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # KPI Cards
    kpis = [
        ("Total Orders",      f"{int(ov['total_orders']):,}",      BLUE,   LIGHT_BLUE),
        ("Total Revenue",     f"R${ov['total_revenue']:,.0f}",      GREEN,  LIGHT_GREEN),
        ("Avg Order Value",   f"R${ov['avg_order_value']:,.2f}",    ORANGE, LIGHT_ORANGE),
        ("Total Customers",   f"{int(ov['total_customers']):,}",    PURPLE, "F5EEF8"),
        ("Avg Review Score",  f"{ov['avg_review_score']:.2f} / 5", GREEN,  LIGHT_GREEN),
        ("Avg Delivery Days", f"{ov['avg_delivery_days']:.1f} days",RED,   LIGHT_RED),
    ]

    # KPI row
    ws.row_dimensions[4].height = 50
    ws.row_dimensions[5].height = 25

    for i, (label, value, color, bg) in enumerate(kpis, 1):
        # Value cell
        val_cell           = ws.cell(row=4, column=i)
        val_cell.value     = value
        val_cell.font      = Font(
            bold=True, size=16,
            color=color, name="Arial"
        )
        val_cell.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        val_cell.fill      = make_header_fill(bg)
        val_cell.border    = make_thick_border()

        # Label cell
        lbl_cell           = ws.cell(row=5, column=i)
        lbl_cell.value     = label
        lbl_cell.font      = Font(
            size=9, color=HEADER_BG, name="Arial"
        )
        lbl_cell.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        lbl_cell.fill      = make_header_fill(bg)
        lbl_cell.border    = make_border()

    # Detailed KPIs table
    ws.cell(row=7, column=1).value = "Detailed Business Metrics"
    ws.cell(row=7, column=1).font  = Font(
        bold=True, size=12,
        color=HEADER_BG, name="Arial"
    )
    ws.row_dimensions[7].height = 22

    metrics = [
        ["Metric",             "Value",
         "Metric",             "Value"],
        ["Total Orders",       f"{int(ov['total_orders']):,}",
         "Total Revenue",      f"R${ov['total_revenue']:,.2f}"],
        ["Total Customers",    f"{int(ov['total_customers']):,}",
         "Total Sellers",      f"{int(ov['total_sellers']):,}"],
        ["Total Products",     f"{int(ov['total_products']):,}",
         "Avg Order Value",    f"R${ov['avg_order_value']:,.2f}"],
        ["Min Order Value",    f"R${ov['min_order_value']:,.2f}",
         "Max Order Value",    f"R${ov['max_order_value']:,.2f}"],
        ["Total Freight",      f"R${ov['total_freight']:,.2f}",
         "Avg Freight",        f"R${ov['avg_freight']:,.2f}"],
        ["Avg Delivery Days",  f"{ov['avg_delivery_days']:.1f}",
         "Late Deliveries",    f"{int(ov['late_deliveries']):,}"],
        ["Avg Review Score",   f"{ov['avg_review_score']:.2f}/5",
         "Late Delivery Rate", f"{ov['late_deliveries']/ov['total_orders']*100:.1f}%"],
    ]

    for i, row_data in enumerate(metrics):
        row_num = 8 + i
        for j, val in enumerate(row_data, 1):
            cell           = ws.cell(row=row_num, column=j)
            cell.value     = val
            cell.border    = make_border()
            cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )
            if i == 0:
                cell.font = Font(
                    bold=True, color=WHITE,
                    size=10, name="Arial"
                )
                cell.fill = make_header_fill(HEADER_BG)
            else:
                alt = i % 2 == 0
                cell.font = Font(size=9, name="Arial")
                cell.fill = make_header_fill(
                    ALT_ROW if alt else WHITE
                )
        ws.row_dimensions[row_num].height = 20

    set_column_widths(ws, [18, 18, 18, 18, 18, 18])
    logger.info("[OK] Overview sheet written")


# ════════════════════════════════════════════════════════════
# SHEET 2: MONTHLY REVENUE
# ════════════════════════════════════════════════════════════

def write_monthly_revenue_sheet(
    wb: Workbook,
    data: dict
) -> None:
    """
    Write monthly revenue sheet with chart.
    """
    logger.info("-- Writing Monthly Revenue sheet --")

    ws       = wb.create_sheet("Monthly Revenue")
    ws.sheet_view.showGridLines = False

    make_title(
        ws, "MONTHLY REVENUE ANALYSIS",
        1, 7, GREEN
    )

    df = data["monthly"]

    write_dataframe(ws, df, start_row=3)

    # Add line chart
    chart       = LineChart()
    chart.title = "Monthly Revenue Trend"
    chart.style = 10
    chart.height= 15
    chart.width = 25

    n_rows = len(df) + 1
    data_ref = Reference(
        ws,
        min_col=4,
        max_col=4,
        min_row=3,
        max_row=3 + n_rows
    )
    cats = Reference(
        ws,
        min_col=3,
        min_row=4,
        max_row=3 + n_rows
    )

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.solidFill = GREEN
    chart.series[0].graphicalProperties.line.width     = 20000

    ws.add_chart(chart, f"A{n_rows + 6}")

    set_column_widths(
        ws, [12, 10, 15, 15, 15, 15, 12]
    )
    logger.info("[OK] Monthly Revenue sheet written")


# ════════════════════════════════════════════════════════════
# SHEET 3: CATEGORY ANALYSIS
# ════════════════════════════════════════════════════════════

def write_category_sheet(
    wb: Workbook,
    data: dict
) -> None:
    """
    Write category analysis sheet with chart.
    """
    logger.info("-- Writing Category sheet --")

    ws       = wb.create_sheet("Category Analysis")
    ws.sheet_view.showGridLines = False

    make_title(
        ws, "PRODUCT CATEGORY ANALYSIS",
        1, 8, ORANGE
    )

    df = data["category"].head(20)
    write_dataframe(ws, df, start_row=3)

    # Add bar chart
    chart       = BarChart()
    chart.type  = "bar"
    chart.title = "Top 20 Categories by Revenue"
    chart.style = 10
    chart.height= 15
    chart.width = 25

    n_rows   = len(df) + 1
    data_ref = Reference(
        ws,
        min_col=4,
        max_col=4,
        min_row=3,
        max_row=3 + n_rows
    )
    cats = Reference(
        ws,
        min_col=1,
        min_row=4,
        max_row=3 + n_rows
    )

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, f"A{n_rows + 6}")

    set_column_widths(
        ws, [25, 10, 12, 15, 12, 15, 12, 12]
    )
    logger.info("[OK] Category sheet written")


# ════════════════════════════════════════════════════════════
# SHEET 4: CUSTOMER ANALYSIS
# ════════════════════════════════════════════════════════════

def write_customer_sheet(
    wb: Workbook,
    data: dict
) -> None:
    """
    Write customer analysis sheet.
    """
    logger.info("-- Writing Customer sheet --")

    ws       = wb.create_sheet("Customer Analysis")
    ws.sheet_view.showGridLines = False

    make_title(
        ws, "CUSTOMER ANALYSIS",
        1, 6, PURPLE
    )

    # Customer segmentation
    ws.cell(row=3, column=1).value = "Customer Segmentation"
    ws.cell(row=3, column=1).font  = Font(
        bold=True, size=11,
        color=PURPLE, name="Arial"
    )

    seg_df = data["segmentation"]
    last_row = write_dataframe(
        ws, seg_df, start_row=4,
        header_color=PURPLE
    )

    # RFM segments
    ws.cell(
        row=last_row + 1, column=1
    ).value = "RFM Segment Analysis"
    ws.cell(
        row=last_row + 1, column=1
    ).font  = Font(
        bold=True, size=11,
        color=BLUE, name="Arial"
    )

    rfm_df = data["rfm"]
    last_row = write_dataframe(
        ws, rfm_df,
        start_row=last_row + 2,
        header_color=BLUE
    )

    # Top customers
    ws.cell(
        row=last_row + 1, column=1
    ).value = "Top 20 Customers by Revenue"
    ws.cell(
        row=last_row + 1, column=1
    ).font  = Font(
        bold=True, size=11,
        color=GREEN, name="Arial"
    )

    cust_df = data["top_customers"]
    write_dataframe(
        ws, cust_df,
        start_row=last_row + 2,
        header_color=GREEN
    )

    # Pie chart for segmentation
    chart = PieChart()
    chart.title  = "Customer Segments"
    chart.style  = 10
    chart.height = 12
    chart.width  = 18

    data_ref = Reference(
        ws, min_col=2, max_col=2,
        min_row=4, max_row=4 + len(seg_df)
    )
    cats = Reference(
        ws, min_col=1,
        min_row=5, max_row=4 + len(seg_df)
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "H4")

    set_column_widths(
        ws, [20, 12, 12, 15, 15, 15]
    )
    logger.info("[OK] Customer sheet written")


# ════════════════════════════════════════════════════════════
# SHEET 5: SALES ANALYSIS
# ════════════════════════════════════════════════════════════

def write_sales_sheet(
    wb: Workbook,
    data: dict
) -> None:
    """
    Write sales analysis sheet.
    """
    logger.info("-- Writing Sales sheet --")

    ws       = wb.create_sheet("Sales Analysis")
    ws.sheet_view.showGridLines = False

    make_title(
        ws, "SALES ANALYSIS",
        1, 6, BLUE
    )

    # Quarterly revenue
    ws.cell(row=3, column=1).value = "Quarterly Revenue"
    ws.cell(row=3, column=1).font  = Font(
        bold=True, size=11,
        color=BLUE, name="Arial"
    )
    qtr_df   = data["quarterly"]
    last_row = write_dataframe(
        ws, qtr_df, start_row=4
    )

    # State revenue
    ws.cell(
        row=last_row + 1, column=1
    ).value = "Revenue by State"
    ws.cell(
        row=last_row + 1, column=1
    ).font  = Font(
        bold=True, size=11,
        color=GREEN, name="Arial"
    )
    state_df = data["state"]
    last_row = write_dataframe(
        ws, state_df,
        start_row=last_row + 2,
        header_color=GREEN
    )

    # Orders by weekday
    ws.cell(
        row=last_row + 1, column=1
    ).value = "Orders by Day of Week"
    ws.cell(
        row=last_row + 1, column=1
    ).font  = Font(
        bold=True, size=11,
        color=ORANGE, name="Arial"
    )
    weekday_df = data["weekday"]
    write_dataframe(
        ws, weekday_df,
        start_row=last_row + 2,
        header_color=ORANGE
    )

    set_column_widths(
        ws, [20, 15, 15, 15, 15, 15]
    )
    logger.info("[OK] Sales sheet written")


# ════════════════════════════════════════════════════════════
# SHEET 6: OPERATIONS
# ════════════════════════════════════════════════════════════

def write_operations_sheet(
    wb: Workbook,
    data: dict
) -> None:
    """
    Write operations analysis sheet.
    """
    logger.info("-- Writing Operations sheet --")

    ws       = wb.create_sheet("Operations")
    ws.sheet_view.showGridLines = False

    make_title(
        ws, "OPERATIONS ANALYSIS",
        1, 7, RED
    )

    # Payment analysis
    ws.cell(row=3, column=1).value = "Payment Analysis"
    ws.cell(row=3, column=1).font  = Font(
        bold=True, size=11,
        color=BLUE, name="Arial"
    )
    pay_df   = data["payment"]
    last_row = write_dataframe(
        ws, pay_df, start_row=4
    )

    # Review analysis
    ws.cell(
        row=last_row + 1, column=1
    ).value = "Review Analysis"
    ws.cell(
        row=last_row + 1, column=1
    ).font  = Font(
        bold=True, size=11,
        color=GREEN, name="Arial"
    )
    rev_df   = data["review"]
    last_row = write_dataframe(
        ws, rev_df,
        start_row=last_row + 2,
        header_color=GREEN
    )

    # Delivery analysis
    ws.cell(
        row=last_row + 1, column=1
    ).value = "Delivery Performance by State"
    ws.cell(
        row=last_row + 1, column=1
    ).font  = Font(
        bold=True, size=11,
        color=ORANGE, name="Arial"
    )
    del_df   = data["delivery"]
    last_row = write_dataframe(
        ws, del_df,
        start_row=last_row + 2,
        header_color=ORANGE
    )

    # Order status
    ws.cell(
        row=last_row + 1, column=1
    ).value = "Order Status Summary"
    ws.cell(
        row=last_row + 1, column=1
    ).font  = Font(
        bold=True, size=11,
        color=RED, name="Arial"
    )
    status_df = data["order_status"]
    write_dataframe(
        ws, status_df,
        start_row=last_row + 2,
        header_color=RED
    )

    set_column_widths(
        ws, [20, 15, 15, 15, 15, 15, 15]
    )
    logger.info("[OK] Operations sheet written")


# ════════════════════════════════════════════════════════════
# SHEET 7: SELLER ANALYSIS
# ════════════════════════════════════════════════════════════

def write_seller_sheet(
    wb: Workbook,
    data: dict
) -> None:
    """
    Write seller analysis sheet.
    """
    logger.info("-- Writing Seller sheet --")

    ws       = wb.create_sheet("Seller Analysis")
    ws.sheet_view.showGridLines = False

    make_title(
        ws, "SELLER PERFORMANCE ANALYSIS",
        1, 8, ORANGE
    )

    df = data["seller"].head(30)
    write_dataframe(ws, df, start_row=3)

    set_column_widths(
        ws, [15, 10, 10, 12, 15, 12, 12, 12]
    )
    logger.info("[OK] Seller sheet written")


# ════════════════════════════════════════════════════════════
# SHEET 8: TOP PRODUCTS
# ════════════════════════════════════════════════════════════

def write_products_sheet(
    wb: Workbook,
    data: dict
) -> None:
    """
    Write top products sheet.
    """
    logger.info("-- Writing Products sheet --")

    ws       = wb.create_sheet("Top Products")
    ws.sheet_view.showGridLines = False

    make_title(
        ws, "TOP PRODUCTS ANALYSIS",
        1, 8, GREEN
    )

    df = data["top_products"]
    write_dataframe(ws, df, start_row=3)

    set_column_widths(
        ws, [15, 20, 10, 12, 15, 12, 12]
    )
    logger.info("[OK] Products sheet written")


# ════════════════════════════════════════════════════════════
# SHEET 9: CUMULATIVE REVENUE
# ════════════════════════════════════════════════════════════

def write_cumulative_sheet(
    wb: Workbook,
    data: dict
) -> None:
    """
    Write cumulative revenue sheet with chart.
    """
    logger.info("-- Writing Cumulative Revenue sheet --")

    ws       = wb.create_sheet("Revenue Trend")
    ws.sheet_view.showGridLines = False

    make_title(
        ws, "CUMULATIVE REVENUE TREND",
        1, 5, BLUE
    )

    df = data["cumulative"]
    write_dataframe(ws, df, start_row=3)

    # Area chart
    chart       = LineChart()
    chart.title = "Cumulative Revenue & Rolling Average"
    chart.style = 10
    chart.height= 15
    chart.width = 28

    n_rows = len(df) + 1

    # Cumulative revenue
    cum_ref = Reference(
        ws, min_col=3, max_col=3,
        min_row=3, max_row=3 + n_rows
    )
    # Rolling avg
    roll_ref = Reference(
        ws, min_col=4, max_col=4,
        min_row=3, max_row=3 + n_rows
    )
    cats = Reference(
        ws, min_col=1,
        min_row=4, max_row=3 + n_rows
    )

    chart.add_data(cum_ref, titles_from_data=True)
    chart.add_data(roll_ref, titles_from_data=True)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties.line.solidFill = BLUE
    chart.series[1].graphicalProperties.line.solidFill = GREEN
    chart.series[1].graphicalProperties.line.dashDot   = "dash"

    ws.add_chart(chart, f"A{n_rows + 6}")

    set_column_widths(ws, [15, 15, 18, 18, 15])
    logger.info("[OK] Cumulative Revenue sheet written")


# ════════════════════════════════════════════════════════════
# SHEET 10: INSIGHTS
# ════════════════════════════════════════════════════════════

def write_insights_sheet(wb: Workbook) -> None:
    """
    Write key insights and recommendations sheet.
    """
    logger.info("-- Writing Insights sheet --")

    ws       = wb.create_sheet("Key Insights")
    ws.sheet_view.showGridLines = False

    make_title(
        ws, "KEY INSIGHTS & RECOMMENDATIONS",
        1, 4, BLUE
    )

    insights = [
        # (Category, Insight, Priority, Color)
        ("Revenue",
         "Total revenue R$19.7M over 2 years with consistent growth",
         "High", GREEN),
        ("Revenue",
         "November 2017 peak due to Black Friday — plan campaigns",
         "High", GREEN),
        ("Revenue",
         "Health & Beauty is top revenue category",
         "Medium", BLUE),
        ("Revenue",
         "Sao Paulo contributes 40%+ of total revenue",
         "High", GREEN),
        ("Customer",
         "97%+ customers are one-time buyers — critical issue",
         "Critical", RED),
        ("Customer",
         "Implement loyalty program to increase repeat purchases",
         "Critical", RED),
        ("Customer",
         "Champions segment needs exclusive rewards",
         "High", GREEN),
        ("Customer",
         "At Risk customers need immediate re-engagement",
         "High", ORANGE),
        ("Delivery",
         "Avg delivery 12 days — improve logistics for remote states",
         "High", ORANGE),
        ("Delivery",
         "Northern states have 25+ day delivery — add warehouses",
         "Critical", RED),
        ("Delivery",
         "8% late delivery rate — penalize consistently late sellers",
         "High", ORANGE),
        ("Seller",
         "Only 9 Platinum sellers — focus on growing Gold tier",
         "High", GREEN),
        ("Seller",
         "Bronze sellers have 98% late delivery — intervene now",
         "Critical", RED),
        ("Seller",
         "Seller training program needed for quality improvement",
         "Medium", BLUE),
        ("Payment",
         "Credit card 74% — offer EMI options to increase AOV",
         "Medium", BLUE),
        ("Payment",
         "Avg 3.7 installments — promote higher value purchases",
         "Low", PURPLE),
    ]

    # Header
    headers = ["Category", "Insight", "Priority", "Action"]
    for j, h in enumerate(headers, 1):
        cell           = ws.cell(row=3, column=j)
        cell.value     = h
        cell.font      = Font(
            bold=True, color=WHITE,
            size=10, name="Arial"
        )
        cell.fill      = make_header_fill(HEADER_BG)
        cell.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        cell.border    = make_border()
    ws.row_dimensions[3].height = 22

    # Data
    for i, (cat, insight, priority, color) in enumerate(
        insights, 4
    ):
        alt = (i - 4) % 2 == 1
        bg  = ALT_ROW if alt else WHITE

        # Category
        c1           = ws.cell(row=i, column=1)
        c1.value     = cat
        c1.font      = Font(
            bold=True, color=color,
            size=9, name="Arial"
        )
        c1.fill      = make_header_fill(bg)
        c1.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        c1.border    = make_border()

        # Insight
        c2           = ws.cell(row=i, column=2)
        c2.value     = insight
        c2.font      = Font(size=9, name="Arial")
        c2.fill      = make_header_fill(bg)
        c2.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )
        c2.border    = make_border()

        # Priority
        priority_colors = {
            "Critical": RED,
            "High"    : ORANGE,
            "Medium"  : BLUE,
            "Low"     : "95A5A6",
        }
        c3           = ws.cell(row=i, column=3)
        c3.value     = priority
        c3.font      = Font(
            bold=True,
            color=priority_colors.get(priority, BLUE),
            size=9, name="Arial"
        )
        c3.fill      = make_header_fill(bg)
        c3.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        c3.border    = make_border()

        # Action
        c4           = ws.cell(row=i, column=4)
        c4.value     = "Review & Implement"
        c4.font      = Font(size=9, name="Arial")
        c4.fill      = make_header_fill(bg)
        c4.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        c4.border    = make_border()

        ws.row_dimensions[i].height = 30

    set_column_widths(ws, [15, 55, 12, 20])
    logger.info("[OK] Insights sheet written")


# ════════════════════════════════════════════════════════════
# MAIN EXCEL REPORT FUNCTION
# ════════════════════════════════════════════════════════════

def generate_excel_report(
    filename: str = "retail_analytics_report.xlsx"
) -> Path:
    """
    Generate complete Excel report with
    multiple sheets and charts.

    Returns:
        Path to generated Excel file
    """
    logger.info("=" * 55)
    logger.info("EXCEL REPORT GENERATION STARTED")
    logger.info("=" * 55)

    create_directories()
    total_start = time.time()

    # Load all data
    logger.info("-- Loading data --")
    data = {
        "overview"    : get_business_overview(),
        "monthly"     : get_monthly_revenue(),
        "quarterly"   : get_quarterly_revenue(),
        "category"    : get_revenue_by_category(),
        "state"       : get_revenue_by_state(),
        "segmentation": get_customer_segmentation(),
        "order_status": get_order_status_summary(),
        "payment"     : get_payment_analysis(),
        "review"      : get_review_analysis(),
        "delivery"    : get_delivery_analysis(),
        "seller"      : get_seller_performance(),
        "rfm"         : get_rfm_segment_summary(),
        "top_customers": get_top_customers(),
        "top_products" : get_top_products(),
        "cumulative"   : get_running_total_revenue(),
        "weekday"      : get_orders_by_weekday(),
    }
    logger.info("[OK] All data loaded")

    # Create workbook
    wb = Workbook()

    # Write all sheets
    write_overview_sheet(wb, data)
    write_monthly_revenue_sheet(wb, data)
    write_category_sheet(wb, data)
    write_customer_sheet(wb, data)
    write_sales_sheet(wb, data)
    write_operations_sheet(wb, data)
    write_seller_sheet(wb, data)
    write_products_sheet(wb, data)
    write_cumulative_sheet(wb, data)
    write_insights_sheet(wb)

    # Save workbook
    output_path = REPORTS_DIR / filename
    wb.save(str(output_path))

    total_elapsed = time.time() - total_start
    file_size     = output_path.stat().st_size / 1024

    print("\n" + "=" * 55)
    print("   EXCEL REPORT SUMMARY")
    print("=" * 55)
    print(f"  File     : {filename}")
    print(f"  Path     : {output_path}")
    print(f"  Size     : {file_size:.1f} KB")
    print(f"  Sheets   : 10")
    print(f"  Time     : {total_elapsed:.2f}s")
    print("-" * 55)
    print("  Sheets:")
    sheets = [
        "Overview", "Monthly Revenue",
        "Category Analysis", "Customer Analysis",
        "Sales Analysis", "Operations",
        "Seller Analysis", "Top Products",
        "Revenue Trend", "Key Insights",
    ]
    for s in sheets:
        print(f"  - {s}")
    print("=" * 55 + "\n")

    logger.info(
        f"[OK] Excel report generated: {output_path}"
    )
    return output_path


# ── Run directly ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    path = generate_excel_report()